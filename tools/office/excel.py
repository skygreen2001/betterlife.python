#!/usr/bin/python
# -*- coding: UTF-8 -*-
#coding=utf-8

# 目标: 读写Excel表格
# 书籍: Python编程快速上手 - 让繁琐工作自动化
# Book: AUTOMATE THE BORING STUFF WITH PYTHON - Practical Programming for Total Beginners.
# [使用python来实现报表自动化](https://blog.csdn.net/coraline_m/article/details/51357185)
# [xlrd、xlwt和openpyxl模块的比较和使用](https://blog.csdn.net/y2Candice/article/details/100581454/)
# [xlrd/xlwt与openpyxl的读写效率比较](https://blog.csdn.net/qq_21391921/article/details/77949645)
# [25.xlrd、xlwt和openpyxl模块的比较和使用](https://www.cnblogs.com/ubuntu1987/p/11933376.html)
#
# 安装需导入的模块
#    > sudo easy_install pip
#    > sudo pip install openpyxl
#    > pip install openpyxl --upgrade
#    > sudo python3 -m pip install --upgrade pip (python3 安装包)
#    > sudo python3 -m pip install openpyxl (python3 安装包)
# 注释: 本示例必须在命令行工具内使用

# 1）xlrd：对xls、xlsx、xlsm文件进行读操作–读操作效率较高，推荐




# 2）xlwt：对xls文件进行写操作–写操作效率较高，但是不能执行xlsx文件




# 3）openpyxl：对xlsx、xlsm文件进行读、写操作–xlsx写操作推荐使用
import os,sys
import datetime
import openpyxl

print(openpyxl.__version__)

path = os.path.realpath(__file__)
if (path):
    path = os.path.split(path)
    path = path[0]
path += '/data/example.xlsx'
print(path)
# os._exit(0)

# 1.新建工作簿：openpyxl.Workbook()
workBook = openpyxl.Workbook()
# 2.在工作簿中新建sheet页：create_sheet()
sheet = workBook.create_sheet("sheet_name")
# 3.向表格中写入数据：cell(i，j，value) - -索引从1计数
sheet = workBook.active # 获得当前活跃的工作页，默认为第一个工作页
sheet.cell(3, 3, "welcome to betterlife python")  # 向单元格（i，j）第i行第j列写入数据value
# 注意：行号和列号都从1开始计数，即（1, 1）为第一行第一列
# 4.保存工作簿：save()
savepath = os.path.realpath(__file__)
if (savepath):
    savepath = os.path.split(savepath)
    savepath = savepath[0]
savepath += '/data/newcreate.xlsx'
workBook.save(savepath)

# 对xlsx文件进行读操作
wb = openpyxl.load_workbook(path)
print(type(wb))
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet3')
print(sheet.title)

sheet['A2'] = datetime.datetime.now()
wb.save(path)

currentSheet = wb.active
print(currentSheet)
print(currentSheet['A1'])
print(currentSheet['A1'].value)

c = currentSheet['B1']
print(c.value)
intro = 'Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value
print(intro)
intro = 'Cell ' + c.coordinate + ' is ' + c.value
print(intro)

cv = currentSheet['C1'].value
print(cv)

print(currentSheet.cell(row=1, column=2))
print(currentSheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, currentSheet.cell(row=i, column=2).value)

count_rows = currentSheet.max_row
count_columns = currentSheet.max_column
print('总行数:{}, 总列数:{}'.format(count_rows, count_columns))

sheets = tuple(currentSheet['A1':'C3'])
print(sheets)

for rowOfCellObjects in currentSheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value) 
    print('--- END OF ROW ---')

columns = currentSheet["B"]
# columns = list(currentSheet.columns)[1]
print(columns)

for cellObj in columns:
    print(cellObj.value)

currentSheet['C8'] = '=SUM(C1:C7)'
wb.save(path)
print(currentSheet['C8'].value)
